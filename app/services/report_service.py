"""
Report generation and export service
"""
import os
import io
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlsxwriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from app.models.asset import Asset
from app.models.client import Client
from app.models.delivery import Delivery
from app.models.report import Report, ReportSchedule, ReportExport


class ReportService:
    """Service for report generation and export"""

    def __init__(self, db: Session):
        self.db = db

    def create_report(
        self,
        name: str,
        report_type: str,
        fields: List[str],
        created_by_admin_id: str,
        description: Optional[str] = None,
        filters: Optional[Dict] = None,
        grouping: Optional[Dict] = None,
        aggregations: Optional[Dict] = None,
        sorting: Optional[Dict] = None,
        is_template: bool = False
    ) -> Report:
        """Create a new custom report"""
        report = Report(
            name=name,
            description=description,
            report_type=report_type,
            fields=fields,
            filters=filters or {},
            grouping=grouping or {},
            aggregations=aggregations or {},
            sorting=sorting or {},
            is_template=is_template,
            created_by_admin_id=created_by_admin_id
        )

        self.db.add(report)
        self.db.commit()
        self.db.refresh(report)

        return report

    def get_report_data(self, report: Report) -> List[Dict]:
        """Execute report query and return data"""
        # Build query based on report type
        if report.report_type == 'leads':
            return self._get_leads_data(report)
        elif report.report_type == 'clients':
            return self._get_clients_data(report)
        elif report.report_type == 'deliveries':
            return self._get_deliveries_data(report)
        elif report.report_type == 'analytics':
            return self._get_analytics_data(report)
        else:
            return []

    def _get_leads_data(self, report: Report) -> List[Dict]:
        """Get leads data for report"""
        query = self.db.query(Asset)

        # Apply filters
        if report.filters:
            for field, condition in report.filters.items():
                if field == 'created_at' and isinstance(condition, dict):
                    if 'from' in condition:
                        query = query.filter(Asset.created_at >= condition['from'])
                    if 'to' in condition:
                        query = query.filter(Asset.created_at <= condition['to'])
                elif field == 'status' and condition:
                    query = query.filter(Asset.status == condition)

        # Apply sorting
        if report.sorting:
            field = report.sorting.get('field', 'created_at')
            order = report.sorting.get('order', 'desc')
            column = getattr(Asset, field, Asset.created_at)
            query = query.order_by(column.desc() if order == 'desc' else column.asc())

        # Execute query
        leads = query.all()

        # Format data based on selected fields
        data = []
        for lead in leads:
            row = {}
            for field in report.fields:
                value = getattr(lead, field, None)
                if isinstance(value, datetime):
                    row[field] = value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    row[field] = str(value) if value is not None else ''
            data.append(row)

        return data

    def _get_clients_data(self, report: Report) -> List[Dict]:
        """Get clients data for report"""
        query = self.db.query(Client)

        # Apply filters
        if report.filters:
            if 'is_active' in report.filters:
                query = query.filter(Client.is_active == report.filters['is_active'])

        # Apply sorting
        if report.sorting:
            field = report.sorting.get('field', 'created_at')
            order = report.sorting.get('order', 'desc')
            column = getattr(Client, field, Client.created_at)
            query = query.order_by(column.desc() if order == 'desc' else column.asc())

        clients = query.all()

        data = []
        for client in clients:
            row = {}
            for field in report.fields:
                value = getattr(client, field, None)
                if isinstance(value, datetime):
                    row[field] = value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    row[field] = str(value) if value is not None else ''
            data.append(row)

        return data

    def _get_deliveries_data(self, report: Report) -> List[Dict]:
        """Get deliveries data for report"""
        query = self.db.query(Delivery)

        # Apply filters
        if report.filters:
            if 'status' in report.filters:
                query = query.filter(Delivery.status == report.filters['status'])
            if 'created_at' in report.filters and isinstance(report.filters['created_at'], dict):
                if 'from' in report.filters['created_at']:
                    query = query.filter(Delivery.timestamp >= report.filters['created_at']['from'])
                if 'to' in report.filters['created_at']:
                    query = query.filter(Delivery.timestamp <= report.filters['created_at']['to'])

        deliveries = query.all()

        data = []
        for delivery in deliveries:
            row = {}
            for field in report.fields:
                value = getattr(delivery, field, None)
                if isinstance(value, datetime):
                    row[field] = value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    row[field] = str(value) if value is not None else ''
            data.append(row)

        return data

    def _get_analytics_data(self, report: Report) -> List[Dict]:
        """Get analytics data for report"""
        # This would aggregate data based on report configuration
        # For now, return summary statistics
        total_leads = self.db.query(func.count(Asset.id)).scalar()
        total_clients = self.db.query(func.count(Client.id)).scalar()
        total_deliveries = self.db.query(func.count(Delivery.id)).scalar()

        return [{
            'metric': 'Total Leads',
            'value': total_leads
        }, {
            'metric': 'Total Clients',
            'value': total_clients
        }, {
            'metric': 'Total Deliveries',
            'value': total_deliveries
        }]

    def export_to_excel(self, report: Report, file_path: str) -> str:
        """Export report to Excel file"""
        data = self.get_report_data(report)

        # Create workbook
        workbook = xlsxwriter.Workbook(file_path)
        worksheet = workbook.add_worksheet(report.name[:31])  # Sheet name limit

        # Define formats
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })

        cell_format = workbook.add_format({
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })

        # Write headers
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers):
                worksheet.write(0, col, header.replace('_', ' ').title(), header_format)
                worksheet.set_column(col, col, 20)  # Set column width

            # Write data
            for row_idx, row_data in enumerate(data, start=1):
                for col_idx, header in enumerate(headers):
                    worksheet.write(row_idx, col_idx, row_data.get(header, ''), cell_format)

        workbook.close()

        return file_path

    def export_to_csv(self, report: Report, file_path: str) -> str:
        """Export report to CSV file"""
        import csv

        data = self.get_report_data(report)

        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            if data:
                headers = list(data[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)

        return file_path

    def export_to_pdf(self, report: Report, file_path: str) -> str:
        """Export report to PDF file"""
        data = self.get_report_data(report)

        # Create PDF
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f497d'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        # Title
        elements.append(Paragraph(report.name, title_style))
        elements.append(Spacer(1, 12))

        # Description
        if report.description:
            elements.append(Paragraph(report.description, styles['Normal']))
            elements.append(Spacer(1, 12))

        # Metadata
        metadata_text = f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} | Records: {len(data)}"
        elements.append(Paragraph(metadata_text, styles['Normal']))
        elements.append(Spacer(1, 20))

        # Table
        if data:
            headers = list(data[0].keys())
            table_data = [[header.replace('_', ' ').title() for header in headers]]

            for row in data:
                table_data.append([str(row.get(header, '')) for header in headers])

            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))

            elements.append(table)

        # Build PDF
        doc.build(elements)

        return file_path

    def create_export(
        self,
        report_id: str,
        export_format: str,
        generated_by_admin_id: str
    ) -> ReportExport:
        """Create a report export record and generate file"""
        report = self.db.query(Report).filter(Report.id == report_id).first()
        if not report:
            raise ValueError("Report not found")

        # Create export directory if it doesn't exist
        export_dir = "/root/leadex-project/exports"
        os.makedirs(export_dir, exist_ok=True)

        # Generate filename
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f"{report.name}_{timestamp}.{export_format}"
        file_path = os.path.join(export_dir, filename)

        # Create export record
        export = ReportExport(
            report_id=report_id,
            export_format=export_format,
            file_path=file_path,
            generated_by_admin_id=generated_by_admin_id,
            status='pending'
        )

        self.db.add(export)
        self.db.commit()

        try:
            # Generate file based on format
            if export_format == 'excel':
                self.export_to_excel(report, file_path)
            elif export_format == 'csv':
                self.export_to_csv(report, file_path)
            elif export_format == 'pdf':
                self.export_to_pdf(report, file_path)
            else:
                raise ValueError(f"Unsupported format: {export_format}")

            # Update export record
            export.status = 'completed'
            export.file_size = os.path.getsize(file_path)
            export.record_count = len(self.get_report_data(report))

        except Exception as e:
            export.status = 'failed'
            export.error_message = str(e)
            raise

        finally:
            self.db.commit()
            self.db.refresh(export)

        return export


# Pre-built report templates
REPORT_TEMPLATES = {
    'lead_performance': {
        'name': 'Lead Performance Report',
        'description': 'Comprehensive lead performance and conversion metrics',
        'report_type': 'leads',
        'fields': ['id', 'mobile', 'name', 'email', 'status', 'utm_source', 'utm_medium', 'utm_campaign', 'created_at'],
        'is_template': True
    },
    'client_activity': {
        'name': 'Client Activity Report',
        'description': 'Client engagement and credit usage statistics',
        'report_type': 'clients',
        'fields': ['id', 'name', 'email', 'credits', 'allocation_percentage', 'is_active', 'created_at'],
        'is_template': True
    },
    'delivery_summary': {
        'name': 'Delivery Summary Report',
        'description': 'Lead delivery status and success rates',
        'report_type': 'deliveries',
        'fields': ['id', 'asset_id', 'client_id', 'channel', 'status', 'retry_count', 'created_at', 'delivered_at'],
        'is_template': True
    }
}
